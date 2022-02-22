# -*- coding: utf-8 -*-
"""
Created on Wed Jan 12 11:16:20 2022

@author: r.thorat
"""
#importing necessary modules
import pandas._libs.tslibs.base
import os
os.getcwd()
import pandas as pd
import re
from nameparser import HumanName
import dash
from dash import dcc
from dash import html
from dash.dependencies import Input, Output, State
import base64
import datetime
import io
from dash import dash_table




external_stylesheets = ['https://codepen.io/chriddyp/pen/bWLwgP.css']

app = dash.Dash(__name__, external_stylesheets=external_stylesheets)

app.layout = html.Div([
    dcc.Upload(
        id='upload-data',
        children=html.Div([
            'Drag and Drop or ',
            html.A('Select Files')
        ]),
        style={
            'width': '100%',
            'height': '60px',
            'lineHeight': '60px',
            'borderWidth': '1px',
            'borderStyle': 'dashed',
            'borderRadius': '5px',
            'textAlign': 'center',
            'margin': '10px'
        },
        # Allow multiple files to be uploaded
        multiple=True
    ),
    html.Div(id='output-data-upload'),
])

def parse_contents(contents, filename, date):
    content_type, content_string = contents.split(',')

    decoded = base64.b64decode(content_string)
    try:
        if 'csv' in filename:
            # Assume that the user uploaded a CSV file
            df = pd.read_csv(
                io.StringIO(decoded.decode('utf-8')))
        elif 'xls' in filename:
            # Assume that the user uploaded an excel file
            df = pd.read_excel(io.BytesIO(decoded))
        df_pv2 = df.groupby(
    ['Dossiernummer', 'Hoofdaanvrager', 'Hoofdaanvrager_Achternaam',
       'Voornaam', 'Geslacht', 'Promotiedatum', 'C_Email',
       'Correspondentietaal', 'HoofdOrganisatie', 'C_Organisatie',
       'C_Postcode', 'C_Plaats', 'Titel', 'Samenvatting', 'Hoofddiscipline',
       'Subdiscipline', 'SubsidieRonde_Naam', 'atl_Ingetrokken',
       'atl_Gehonoreerd'],
)['Woord'].agg(', '.join).reset_index(name='Trefwoorden')

        # # Stap 2 Edit information in terms of 'Proposal Upload Example.xlsx' from expert look up
        #2.a Formulating information for 'Proposal Details' worksheet
        df_PD= pd.DataFrame(columns = ['No', 'Grant No.', 'Council', 'Proposal Title', 'Abstract', 'Specific Aims', 'Keywords'])
        
        
        
        #Changing names of the columns
        df_PD['Grant No.']=df_pv2['Dossiernummer']
        df_PD['No'] = df_pv2['Dossiernummer'].str[-3:]
        df_PD['Council']='NWO'
        df_PD['Proposal Title']=df_pv2['Titel']
        df_PD['Abstract']=df_pv2['Samenvatting']
        df_PD['Keywords']=df_pv2['Trefwoorden']
        #check
        df_PD.head()
        # 2.b Formulate information for 'Proposal Applicants' worksheet, main applicant
        df_PA= pd.DataFrame(columns = ['Proposal Number', 'Grant No.', 'Last Name', 'First Name',
                                       'Scopus Author ID', 'OrcId', 'Email', 'Affiliation', 'Country', 'Role', 'Is Principal Investigator?'])
        #Changing names of the columns
       
        df_PA['Grant No.']=df_pv2['Dossiernummer']
        df_PA['Proposal Number'] = df_pv2['Dossiernummer'].str[-3:]
        df_PA['Last Name']=df_pv2['Hoofdaanvrager_Achternaam']
        df_PA['First Name']=df_pv2['Voornaam']
        df_PA['Email']=df_pv2['C_Email']
        df_PA['Affiliation']=df_pv2['HoofdOrganisatie']
        df_PA['Country']='Netherlands'
        df_PA['Role']='Principal Investigator'
        df_PA['Is Principal Investigator?']='Yes'
        #The openpyxl.utils.dataframe.dataframe_to_rows() function provides a simple way to work with Pandas Dataframes:

        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl import Workbook
        
        
        wb = Workbook()
        ws = wb.active
        #ws = wb.create_sheet("Samenvatting", 0) 
        ws.title='Proposal Details'
        
        for r in dataframe_to_rows(df_PD, index=False, header=True):
            ws.append(r)
        
        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'
        
        #writing proposal applicants
        ws1 = wb.create_sheet("Proposal Applicants", 1) 
        for r in dataframe_to_rows(df_PA, index=False, header=True):
            ws1.append(r)
            
        #saving the excel workbook
        Exportfile_name=('Proposal info to feed in EL.xlsx ')
        wb.save(Exportfile_name)


    except Exception as e:
        print(e)
        return html.Div([
            'There was an error processing this file.'
        ])

    return html.Div([
        html.H5(filename),
        html.H6(datetime.datetime.fromtimestamp(date)),

        dash_table.DataTable(
            data=df.to_dict('records'),
            columns=[{'name': i, 'id': i} for i in df.columns]
        ),

        html.Hr(),  # horizontal line

        # For debugging, display the raw contents provided by the web browser
        html.Div('Raw Content'),
        html.Pre(contents[0:2] + '...', style={
            'whiteSpace': 'pre-wrap',
            'wordBreak': 'break-all'
        })
    ])

@app.callback(Output('output-data-upload', 'children'),
              Input('upload-data', 'contents'),
              State('upload-data', 'filename'),
              State('upload-data', 'last_modified'))
def update_output(list_of_contents, list_of_names, list_of_dates):
    if list_of_contents is not None:
        children = [
            parse_contents(c, n, d) for c, n, d in
            zip(list_of_contents, list_of_names, list_of_dates)]
        return children

if __name__ == '__main__':
    app.run_server(debug=True)



